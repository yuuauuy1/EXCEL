#必要ライブラリのインポート
import openpyxl 
import csv
from fractions import Fraction  #有理数(分数で表現できるような値)を扱うためのモジュール

#--------------------ファイルを開く-----------------------------------
ing_wb = openpyxl.load_workbook("recipe_ingredients.xlsx") #レシピの材料と量の一覧表
ing_ws = ing_wb.worksheets[0] #ワークシート１枚目を選択

index_wb = openpyxl.load_workbook("recipe_index.xlsx") #レシピ名、ＵＲＬ、材料、人数分の表一覧
index_ws = index_wb.worksheets[0]

#csvファイルからUIとなるword_listを作る
word_list = []
def create_wordlist():
    with open("word_list.csv" , "r", encoding="utf-8") as f:
        reader = csv.reader(f)        
        for row in reader:
            word_list.append(row[0])
  
#index.レシピ名.URL.人数.レシピ名と材料が一緒になった列のリストを作る
index_name_lists=[]
def create_index_namelist(): 
    
    for row in index_ws.iter_rows(min_row =2, min_col=1, max_col=6):
        index_name =[]
        index_name.append(row[0].value) #index列
        index_name.append(row[1].value) #レシピ名
        index_name.append(row[2].value) #URL
        index_name.append(row[3].value) #人数
        index_name.append(row[5].value) #レシピ名と材料が一緒になった列　

        index_name_lists.append(index_name)


#index.材料.量が一緒になったリストを作る
index_ing_lists=[]
def create_index_inglist():   
    
    for row in ing_ws.iter_rows(min_row =2, min_col=1, max_col=6):
        index_ing =[]
        index_ing.append(row[0].value) #index列
        index_ing.append(row[1].value) #材料
        index_ing.append(row[2].value) #量
        index_ing.append(row[3].value) #量(数字のみ)
        index_ing.append(row[4].value) #単位
        index_ing.append(row[5].value) #材料+単位
        index_ing_lists.append(index_ing)


# 検索文字が正しく入力されているか判定する
def input_word():  
    word = str(input(word_list[0]))
    word_result =  word.isalpha() #ひらがな／カタカナ／漢字が文字判定(True)される。数字は除外(False)
    if word_result :
        search_word(word)
    else:
        print(word_list[1]) #日本語以外の文字が入力されました      
        input_word()  
    

#検索文字が一致するセルの座標をrecipe_indexファイルから検索
def search_word(word):    
    title_list = [] #検索件数を把握するためにtitle_list にtitle格納する
    for index_name in index_name_lists: 
         # 検索ワードと一致した時        
        if word in index_name[4]:
          
            # index と URLを取得する
            index = index_name[0]
            title = index_name[1]
            url = index_name[2]
            people_n = index_name[3]
            title_list.append(title)

            message(index, title, url, people_n)

            for index_ing in index_ing_lists:   #リストとして持っておく
                if str(index) in str(index_ing[0]):
                    ingredient = index_ing[1]
                    amount =  index_ing[2]           
                    print("{"+ str(ingredient)+"："+str(amount)+"}")

    print( str( len(title_list))+ "件のレシピが見つかりました。" )

    if len(title_list) == 0: # 検索結果無い場合、検索ワード入力に戻る
        print(word_list[2]) #検索ワードの文字表記を変えると検索結果変わるかも・・・
        input_word()
    
    else: # 検索結果がある場合、レシピのindex選択に進む
        input_index()

def message(index, title, url, people_n):
    print("+++++++++++++++++++++++++++++++++++++++++")
    print("index : "+ str(index))
    print("レシピ名 : "+ title)
    print(str(people_n)+ " 人分のレシピです")
    print("URL : "+ url)
    
index_list =[] # レシピindexのリスト    
def input_index(): # 買い物リストに入れるレシピの index が 正しい数字か判定
    recipe_index = input( word_list[3] ) #買い物リストに入れるレシピの　indexを入力してください
    index_result = recipe_index.isdigit() # 正の整数か判定
    
    if index_result:
        if 0 <= int(recipe_index) <= 2031:
            index_list.append(int(recipe_index))
            end()        
        else:
            print(word_list[4]) #indexにない数字が入力されました
            input_index()
    else:
        print(word_list[5]) #数字以外の文字が入力されました
        input_index()

def end(): # レシピ検索継続するか 終了するか入力
    input_n = input(word_list[6]) #まだレシピ検索しますか？Yes:1,No:0
    n_result = input_n.isdigit() # 正の整数か判定
    if n_result:
        
        if int(input_n) == 1:
            input_word()
        
        elif int(input_n) == 0:
            print("レシピ選択を終わります。 現在" + str(len(index_list)) + "個のレシピが選択されています。")
            print("選択されたレシピ index : " + str(index_list))
            print(word_list[8]) #それでは買い物リストを作りましょう
        else:
            print(word_list[7]) #1 or 0で選択してください
            end()
    else:
        print(word_list[5]) #数字以外が入力されました
        end()

#選択したレシピの材料を何人分のお買い物リストとしたいか選択
def choose_n_person():   
    count = 1
    dic_lists = []
    for index in index_list:
    
        for index_name in index_name_lists:

            if int(index) == index_name[0] : 
                number_of_people = index_name[3] #何人分のレシピか
                title = index_name[1] #レシピの名前
                print("*********************************")
                print(str(count) + "番目, レシピNo : " +str(index)+ " , こちらは " +str(number_of_people) + " 人分のレシピです")
                print("レシピ名 「 "+ title+" 」")
                                
                def calc_fractions():
                    print(word_list[10])#何人分の材料を買いたいですか？
                    human_n = input(word_list[9])  #ユーザーが作りたい人数を入力                           
                    human_n_result = human_n.isdigit() #ユーザーが作りたい人数が整数であるか判定
                               
                    if human_n_result: 
                        ingredients =[]
                        amount_list =[]   
                        for index_ing in index_ing_lists:
        
                            if int(index) == index_ing[0]:
                                ingredients.append(index_ing[5])
                                num = Fraction(index_ing[3])
                                new_num = num * Fraction(int(human_n), int(number_of_people)) #欲しい人数の量に変換
                                amount_list.append(new_num)
                                dic= dict(zip(ingredients, amount_list)) 

                        dic_lists.append(dic) # 辞書同士を連結させるために、辞書のリストを作る
                    else:
                        print(word_list[5]) #【！】 数字以外の文字が入力されました。
                        calc_fractions()
                calc_fractions()
        count += 1
    return dic_lists

#材料と量一覧を取得する
new_dic = {}
def get_ingredirnts(dic_lists):
    for dic in dic_lists:                      
        for key in dic.keys() : 

            if key in new_dic: # dicとnew_dicに同じkeyがある場合はvalueを足す
                num_1 = Fraction(new_dic.get(key) or 0) #分数の分量があるのでFraction型に変換する      
                num_2 = Fraction(dic.get(key) or 0) #分量Noneがある可能性あるので　「or 0」 を入力      
                sum = num_1 + num_2          
                new_dic[key] = sum                
            else: # 同じkeyがない場合は　dicのvalueをそのまま採用
                new_dic[key] = dic[key]
                
#--------------実行------------------------
create_wordlist()
create_index_namelist()
create_index_inglist()
input_word()
lists = choose_n_person()
get_ingredirnts(lists)

#お買い物リスト表示
print(word_list[11]) #----買い物リスト---
print(word_list[12])#調味料は０と表示されますが、１つあれば良いという意味です！
print(word_list[13])#（）の中は単位です
for key, value in new_dic.items():
    
    print(key + " : " +  str(value))